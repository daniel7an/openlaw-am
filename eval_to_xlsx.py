"""Bundle eval result JSONs into one presentable Excel workbook.

One column per run on the Summary sheet (metrics as rows, so the k-sweep reads
side by side), a By-difficulty sheet, and one per-question sheet per run.
Summaries are recomputed from the stored rows via eval.summarize, so a scorer
fix applies to old files without re-running any model.

Usage:
    uv run python eval_to_xlsx.py results/run1.json results/run2.json
    uv run python eval_to_xlsx.py results/*.json --out results/eval_scores.xlsx
"""
import json
import sys
from pathlib import Path

import eval as ev

PCT, NUM = "0%", "0.0"


def run_label(meta: dict, metas: list[dict]) -> str:
    label = f"k={meta.get('top_k', '?')}"
    if len({m.get("answer_model") for m in metas}) > 1:
        label += f" {meta.get('answer_model', '').split('/')[-1]}"
    return label


def load(path: str) -> dict:
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    payload["summary"] = ev.summarize(payload["rows"])
    return payload


def own_hit_key(meta: dict, rows: list[dict]) -> str:
    k = f"hit@{meta.get('top_k')}"
    return k if rows and k in rows[0]["retrieval"] else "hit@8"


def style_header(ws, widths: dict[str, int]) -> None:
    from openpyxl.styles import Font

    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "B2"


def main() -> None:
    from openpyxl import Workbook

    argv = sys.argv[1:]
    out = Path("results/eval_scores.xlsx")
    if "--out" in argv:
        i = argv.index("--out")
        out = Path(argv[i + 1])
        del argv[i : i + 2]
    args = argv
    if not args:
        sys.exit("usage: uv run python eval_to_xlsx.py results/run1.json ... [--out x.xlsx]")

    payloads = [load(a) for a in args]
    metas = [p["meta"] for p in payloads]
    labels = [run_label(m, metas) for m in metas]

    wb = Workbook()

    # --- Summary: metrics as rows, one column per run -------------------------
    ws = wb.active
    ws.title = "Summary"
    hit_keys = sorted(
        {k for p in payloads for k in p["summary"]["retrieval"]},
        key=lambda s: int(s.split("@")[1]),
    )
    info = [
        ("scope", lambda p: p["meta"].get("scope")),
        ("questions (n)", lambda p: p["summary"]["n"]),
        ("answer model", lambda p: p["meta"].get("answer_model")),
        ("judge model", lambda p: p["meta"].get("judge_model")),
        ("search mode / alpha", lambda p: f"{p['meta'].get('search_mode')} / {p['meta'].get('alpha')}"),
        ("top_k", lambda p: p["meta"].get("top_k")),
    ]
    pct_metrics = [
        *[(k, lambda p, kk=k: p["summary"]["retrieval"].get(kk)) for k in hit_keys],
        ("citation recall (vs gold)", lambda p: p["summary"]["citation_recall_mean"]),
        ("citation precision (vs gold)", lambda p: p["summary"]["citation_precision_vs_gold_mean"]),
        ("exact citation match", lambda p: p["summary"]["exact_citation_match_rate"]),
        ("hallucinated citation rate", lambda p: p["summary"]["hallucinated_citation_rate_mean"]),
        ("judge accuracy (1/0)", lambda p: p["summary"]["judge_fully_correct_rate"]),
    ]
    num_metrics = [
        ("questions judged", lambda p: p["summary"]["n_judged"]),
        ("citations per answer", lambda p: p["summary"]["citations_per_answer_mean"]),
        ("tokens per question", lambda p: p["summary"]["tokens_mean"]),
        ("latency s per question", lambda p: p["summary"]["latency_s_mean"]),
    ]

    ws.append(["metric", *labels])
    for name, get in info:
        ws.append([name, *[get(p) for p in payloads]])
    for name, get in pct_metrics:
        ws.append([name, *[get(p) for p in payloads]])
        for cell in ws[ws.max_row][1:]:
            cell.number_format = PCT
    for name, get in num_metrics:
        ws.append([name, *[get(p) for p in payloads]])
        for cell in ws[ws.max_row][1:]:
            cell.number_format = NUM
    style_header(ws, {"A": 30, **{chr(66 + i): 22 for i in range(len(payloads))}})

    # --- By difficulty --------------------------------------------------------
    ws = wb.create_sheet("By difficulty")
    ws.append(["run", "difficulty", "n", "hit@5", "citation recall", "exact match",
               "judge mean", "n judged"])
    for label, p in zip(labels, payloads):
        for diff, d in p["summary"]["by_difficulty"].items():
            ws.append([label, diff, d["n"], d["hit@5"], d["citation_recall"],
                       d["exact_citation_match"], d["judge_correctness_mean"], d["n_judged"]])
            for cell in ws[ws.max_row][3:7]:
                cell.number_format = PCT
    style_header(ws, {"A": 14, "B": 10, "D": 10, "E": 14, "F": 12, "G": 12})

    # --- Per-question sheet per run -------------------------------------------
    for label, p in zip(labels, payloads):
        ws = wb.create_sheet(f"Q {label}"[:31])
        hk = own_hit_key(p["meta"], p["rows"])
        ws.append(["id", "difficulty", "question (EN)", "expected", "cited", hk,
                   "citation recall", "exact", "hallucinated", "judge", "judge rationale",
                   "tokens", "latency s"])
        for r in p["rows"]:
            c, j = r["citations"], r["judge"]
            ws.append([
                r["id"], r["difficulty"], r.get("question_english", ""),
                ", ".join(c["expected_articles"]), ", ".join(c["cited_articles"]),
                int(bool(r["retrieval"].get(hk))), c["citation_recall"],
                int(c["exact_citation_match"]), ", ".join(c["hallucinated_citations"]),
                j.get("correctness"), (j.get("rationale") or "")[:250],
                r["total_tokens"], round(r["latency_s"], 1),
            ])
            ws[ws.max_row][6].number_format = PCT
        style_header(ws, {"A": 8, "C": 46, "D": 14, "E": 18, "G": 12, "K": 60})

    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    print(f"wrote {out}  ({len(payloads)} runs: {', '.join(labels)})")


if __name__ == "__main__":
    main()
